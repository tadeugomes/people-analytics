#!/usr/bin/env python3
"""
Inject Excel-only dynamic-formula diversity sheets into an existing workbook (e.g., .xlsm),
without removing existing content. Sheets are added with the suffix _XL to avoid conflicts.
 
Usage:
  python inject_excel_only_sheets.py <workbook_path>
"""

import sys
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import BarChart, Reference
from openpyxl.worksheet.table import Table, TableStyleInfo


def norm(s: str) -> str:
    import unicodedata
    s2 = unicodedata.normalize('NFKD', str(s)).encode('ASCII', 'ignore').decode('ASCII')
    return ''.join(ch if ch.isalnum() else '_' for ch in s2.lower()).strip('_')


def find_col(cols, candidates):
    mapping = {norm(c): c for c in cols}
    for k in mapping:
        for cand in candidates:
            if norm(cand) == k or norm(cand) in k:
                return mapping[k]
    return None


def ensure_table(ws, name='TBL_DADOS'):
    # Add a table on DADOS_BRUTOS if not present
    if not ws.tables:
        max_row = ws.max_row
        max_col = ws.max_column
        ref = f"A1:{get_column_letter(max_col)}{max_row}"
        tbl = Table(displayName=name, ref=ref)
        style = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False, showLastColumn=False,
                               showRowStripes=True, showColumnStripes=False)
        tbl.tableStyleInfo = style
        ws.add_table(tbl)


def apply_threshold_format(ws, idx_col_letter='C'):
    from openpyxl.formatting.rule import FormulaRule
    red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    yellow = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    max_row = 2000
    last_col = get_column_letter(ws.max_column if ws.max_column > 5 else 5)
    ref_range = f"$A$2:${last_col}${max_row}"
    ws.conditional_formatting.add(ref_range, FormulaRule(formula=[f"=${idx_col_letter}2<0.6"], fill=red))
    ws.conditional_formatting.add(ref_range, FormulaRule(formula=[f"=AND(${idx_col_letter}2>=0.6,${idx_col_letter}2<0.8)"], fill=yellow))
    ws.conditional_formatting.add(ref_range, FormulaRule(formula=[f"=${idx_col_letter}2>=0.8"], fill=green))


def sheet_name_for(prefix, var):
    return f"{prefix}_{var}"[:31]


def build_gender_sheet(wb, table_name, var_col, gender_col):
    title = sheet_name_for("3A_XL_GENERO", var_col)
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    ws['A1'].value = 'Categoria'; ws['A1'].font = Font(bold=True)
    ws['B1'].value = 'Total'; ws['B1'].font = Font(bold=True)
    ws['C1'].value = 'Índice de Simpson (Gênero)'; ws['C1'].font = Font(bold=True)
    ws['D1'].value = 'Índice de Shannon (Gênero)'; ws['D1'].font = Font(bold=True)
    ws['E1'].value = 'Predominante (Gênero)'; ws['E1'].font = Font(bold=True)
    ws['A2'].value = f"=UNIQUE({table_name}[{var_col}])"
    ws['B2'].value = f"=IF(A2=\"\",\"\",ROWS(FILTER({table_name}[{gender_col}], {table_name}[{var_col}]=A2)))"
    ws['C2'].value = (
        f"=IF(A2=\"\",\"\",LET(arr,FILTER({table_name}[{gender_col}],{table_name}[{var_col}]=A2),"
        f"u,UNIQUE(arr),p,COUNTIF(arr,u)/ROWS(arr),1-SUMPRODUCT(p^2)))"
    )
    ws['D2'].value = (
        f"=IF(A2=\"\",\"\",LET(arr,FILTER({table_name}[{gender_col}],{table_name}[{var_col}]=A2),"
        f"u,UNIQUE(arr),p,COUNTIF(arr,u)/ROWS(arr),pp,FILTER(p,p>0),-SUMPRODUCT(pp*LN(pp))))"
    )
    ws['E2'].value = (
        f"=IF(A2=\"\",\"\",LET(arr,FILTER({table_name}[{gender_col}],{table_name}[{var_col}]=A2),"
        f"u,UNIQUE(arr),c,COUNTIF(arr,u),INDEX(SORTBY(u,c,-1),1)))"
    )
    apply_threshold_format(ws, 'C')
    return ws


def build_race_sheet(wb, table_name, var_col, race_col):
    title = sheet_name_for("3C_XL_RACA", var_col)
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    ws['A1'].value = 'Categoria'; ws['A1'].font = Font(bold=True)
    ws['B1'].value = 'Total'; ws['B1'].font = Font(bold=True)
    ws['C1'].value = 'Índice de Simpson (Raça)'; ws['C1'].font = Font(bold=True)
    ws['D1'].value = 'Índice de Shannon (Raça)'; ws['D1'].font = Font(bold=True)
    ws['E1'].value = 'Predominante (Raça)'; ws['E1'].font = Font(bold=True)
    ws['A2'].value = f"=UNIQUE({table_name}[{var_col}])"
    ws['B2'].value = f"=IF(A2=\"\",\"\",ROWS(FILTER({table_name}[{race_col}], {table_name}[{var_col}]=A2)))"
    ws['C2'].value = (
        f"=IF(A2=\"\",\"\",LET(arr,FILTER({table_name}[{race_col}],{table_name}[{var_col}]=A2),"
        f"u,UNIQUE(arr),p,COUNTIF(arr,u)/ROWS(arr),1-SUMPRODUCT(p^2)))"
    )
    ws['D2'].value = (
        f"=IF(A2=\"\",\"\",LET(arr,FILTER({table_name}[{race_col}],{table_name}[{var_col}]=A2),"
        f"u,UNIQUE(arr),p,COUNTIF(arr,u)/ROWS(arr),pp,FILTER(p,p>0),-SUMPRODUCT(pp*LN(pp))))"
    )
    ws['E2'].value = (
        f"=IF(A2=\"\",\"\",LET(arr,FILTER({table_name}[{race_col}],{table_name}[{var_col}]=A2),"
        f"u,UNIQUE(arr),c,COUNTIF(arr,u),INDEX(SORTBY(u,c,-1),1)))"
    )
    apply_threshold_format(ws, 'C')
    return ws


def build_gender_summary(wb, variables):
    title = '3B_XL_RESUMO_DIVERSIDADE_GENERO'
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    headers = ['Variável','Categorias (Qtd)','Total (registros)','Índice de Simpson (Gênero) — média ponderada','Baixa (<0,6)','Moderada (0,6–<0,8)','Alta (>=0,8)','Sinal']
    for j,h in enumerate(headers, start=1): ws.cell(row=1,column=j,value=h).font=Font(bold=True)
    for i,var in enumerate(variables, start=2):
        ws.cell(row=i, column=1, value=var)
        sname = sheet_name_for('3A_XL_GENERO', var)
        cat = f"'{sname}'!A2:A1048576"; tot = f"'{sname}'!B2:B1048576"; sim = f"'{sname}'!C2:C1048576"
        ws.cell(row=i, column=2, value=f"=IFERROR(COUNTA(FILTER({cat},{cat}<>\"\")),0)")
        ws.cell(row=i, column=3, value=f"=IFERROR(SUM(FILTER({tot},{cat}<>\"\")),0)")
        ws.cell(row=i, column=4, value=f"=IFERROR(SUMPRODUCT(FILTER({sim},{cat}<>\"\"),FILTER({tot},{cat}<>\"\"))/SUM(FILTER({tot},{cat}<>\"\")),\"\")")
        ws.cell(row=i, column=5, value=f"=LET(c,FILTER({sim},{cat}<>\"\"),IF(c=\"\",0,SUMPRODUCT(--(c<0.6))))")
        ws.cell(row=i, column=6, value=f"=LET(c,FILTER({sim},{cat}<>\"\"),IF(c=\"\",0,SUMPRODUCT(--(c>=0.6),--(c<0.8))))")
        ws.cell(row=i, column=7, value=f"=LET(c,FILTER({sim},{cat}<>\"\"),IF(c=\"\",0,SUMPRODUCT(--(c>=0.8))))")
        ws.cell(row=i, column=8, value=f"=IF(D{i}<0.6,\"🔴\",IF(D{i}<0.8,\"🟡\",\"🟢\"))")
    apply_threshold_format(ws, 'D')


def build_race_summary(wb, variables):
    title = '3D_XL_RESUMO_DIVERSIDADE_RACA'
    if title in wb.sheetnames:
        del wb[title]
    ws = wb.create_sheet(title)
    headers = ['Variável','Categorias (Qtd)','Total (registros)','Índice de Simpson (Raça) — média ponderada','Baixa (<0,6)','Moderada (0,6–<0,8)','Alta (>=0,8)','Sinal']
    for j,h in enumerate(headers, start=1): ws.cell(row=1,column=j,value=h).font=Font(bold=True)
    for i,var in enumerate(variables, start=2):
        ws.cell(row=i, column=1, value=var)
        sname = sheet_name_for('3C_XL_RACA', var)
        cat = f"'{sname}'!A2:A1048576"; tot = f"'{sname}'!B2:B1048576"; sim = f"'{sname}'!C2:C1048576"
        ws.cell(row=i, column=2, value=f"=IFERROR(COUNTA(FILTER({cat},{cat}<>\"\")),0)")
        ws.cell(row=i, column=3, value=f"=IFERROR(SUM(FILTER({tot},{cat}<>\"\")),0)")
        ws.cell(row=i, column=4, value=f"=IFERROR(SUMPRODUCT(FILTER({sim},{cat}<>\"\"),FILTER({tot},{cat}<>\"\"))/SUM(FILTER({tot},{cat}<>\"\")),\"\")")
        ws.cell(row=i, column=5, value=f"=LET(c,FILTER({sim},{cat}<>\"\"),IF(c=\"\",0,SUMPRODUCT(--(c<0.6))))")
        ws.cell(row=i, column=6, value=f"=LET(c,FILTER({sim},{cat}<>\"\"),IF(c=\"\",0,SUMPRODUCT(--(c>=0.6),--(c<0.8))))")
        ws.cell(row=i, column=7, value=f"=LET(c,FILTER({sim},{cat}<>\"\"),IF(c=\"\",0,SUMPRODUCT(--(c>=0.8))))")
        ws.cell(row=i, column=8, value=f"=IF(D{i}<0.6,\"🔴\",IF(D{i}<0.8,\"🟡\",\"🟢\"))")
    apply_threshold_format(ws, 'D')


def build_visuals(wb, variables, has_gender, has_race):
    ws = wb.create_sheet('5_XL_VISUALIZACOES')
    ws['A1'].value = 'Visualizações (Excel-only)'
    ws['A1'].font = Font(bold=True, size=14)
    row = 3
    for var in variables:
        if has_gender:
            sname = sheet_name_for('3A_XL_GENERO', var)
            ws.cell(row=row, column=1, value=f"{var} — Gênero").font = Font(bold=True)
            chart = BarChart(); chart.title = f"Índice de Simpson (Gênero) — {var}"; chart.y_axis.title = 'Índice'; chart.x_axis.title = 'Categoria'
            cats = Reference(wb[sname], min_col=1, min_row=2, max_row=200)
            data = Reference(wb[sname], min_col=3, min_row=1, max_row=200)
            chart.add_data(data, titles_from_data=True); chart.set_categories(cats); chart.height=10; chart.width=24
            ws.add_chart(chart, f"A{row+1}"); row += 22
        if has_race:
            sname = sheet_name_for('3C_XL_RACA', var)
            ws.cell(row=row, column=1, value=f"{var} — Raça").font = Font(bold=True)
            chart = BarChart(); chart.title = f"Índice de Simpson (Raça) — {var}"; chart.y_axis.title = 'Índice'; chart.x_axis.title = 'Categoria'
            cats = Reference(wb[sname], min_col=1, min_row=2, max_row=200)
            data = Reference(wb[sname], min_col=3, min_row=1, max_row=200)
            chart.add_data(data, titles_from_data=True); chart.set_categories(cats); chart.height=10; chart.width=24
            ws.add_chart(chart, f"A{row+1}"); row += 22


def main():
    if len(sys.argv) < 2:
        print("Usage: inject_excel_only_sheets.py <workbook_path>")
        sys.exit(1)
    wb_path = sys.argv[1]
    keep_vba = wb_path.lower().endswith('.xlsm')
    wb = load_workbook(wb_path, keep_vba=keep_vba)

    if 'DADOS_BRUTOS' not in wb.sheetnames:
        print("✗ DADOS_BRUTOS not found in workbook")
        sys.exit(2)
    ws_data = wb['DADOS_BRUTOS']
    ensure_table(ws_data, 'TBL_DADOS')

    # Read DADOS_BRUTOS via pandas to infer columns
    df = pd.read_excel(wb_path, sheet_name='DADOS_BRUTOS')
    gender_col = find_col(df.columns, ['genero','gênero','sexo','gender'])
    race_col = find_col(df.columns, ['raca','raça','raca_cor','cor','race','etnia','ethnicity'])
    # Choose up to 5 categorical variables different from gender/race
    exclude = set([c for c in [gender_col, race_col] if c])
    cat_candidates = [c for c in df.columns if c not in exclude and df[c].nunique(dropna=True) <= 20]
    variables = cat_candidates[:5]

    # Build per-variable sheets with formulas referencing TBL_DADOS
    for var in variables:
        if gender_col:
            build_gender_sheet(wb, 'TBL_DADOS', var, gender_col)
        if race_col:
            build_race_sheet(wb, 'TBL_DADOS', var, race_col)

    # Summaries
    if gender_col:
        build_gender_summary(wb, variables)
    if race_col:
        build_race_summary(wb, variables)

    # Visuals
    build_visuals(wb, variables, bool(gender_col), bool(race_col))

    wb.save(wb_path)
    print(f"✓ Injected Excel-only sheets into: {wb_path}")


if __name__ == '__main__':
    main()
