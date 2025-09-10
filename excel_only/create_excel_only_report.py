#!/usr/bin/env python3
"""
Create an Excel-only (no Python needed to refresh) diversity report using dynamic Excel formulas.
Input: CSV path and output XLSX path
"""
import sys
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import BarChart, Reference
from core.utils import load_config, get_thresholds
 

GENDER_NAMES = [
    'genero', 'gênero', 'sexo', 'gender'
]
RACE_NAMES = [
    'raca', 'raça', 'raca_cor', 'cor', 'race', 'etnia', 'ethnicity'
]


def norm(s: str) -> str:
    import unicodedata
    s2 = unicodedata.normalize('NFKD', str(s)).encode('ASCII', 'ignore').decode('ASCII')
    return ''.join(ch if ch.isalnum() else '_' for ch in s2.lower()).strip('_')


def find_col(df: pd.DataFrame, candidates) -> str | None:
    mapping = {norm(c): c for c in df.columns}
    for k in mapping:
        for cand in candidates:
            if norm(cand) == k or norm(cand) in k:
                return mapping[k]
    return None


def auto_categorical_cols(df: pd.DataFrame, exclude: set[str]) -> list[str]:
    cols = []
    for c in df.columns:
        if c in exclude:
            continue
        if df[c].nunique(dropna=True) <= 20 and df[c].dtype != float:
            cols.append(c)
    return cols[:5]  # limit to 5 to keep workbook compact


def add_table(ws, max_row, max_col, name='TBL_DADOS'):
    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    tbl = Table(displayName=name, ref=ref)
    style = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False,
                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    tbl.tableStyleInfo = style
    ws.add_table(tbl)


def add_home(ws):
    ws['A1'].value = 'Relatório Excel-Only (Fórmulas Dinâmicas)'
    ws['A1'].font = Font(bold=True, size=16)
    ws['A3'].value = 'Como usar: edite DADOS_BRUTOS; as abas recalculam automaticamente.'
    ws['A5'].value = 'Legenda de cores (Índice de Simpson)'
    ws['A5'].font = Font(bold=True)
    r = 6
    ws[f'A{r}'].value = '< 0,6 = Baixa diversidade'
    ws[f'A{r}'].fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    ws[f'A{r+1}'].value = '0,6 – <0,8 = Diversidade moderada'
    ws[f'A{r+1}'].fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    ws[f'A{r+2}'].value = '≥ 0,8 = Alta diversidade'
    ws[f'A{r+2}'].fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')


def build_gender_sheet(wb, table_name, var_col, gender_col):
    title = f"3A_GENERO_{var_col}"
    ws = wb.create_sheet(title[:31])
    ws['A1'].value = 'Categoria'
    ws['B1'].value = 'Total'
    ws['C1'].value = 'Índice de Simpson (Gênero)'
    ws['D1'].value = 'Índice de Shannon (Gênero)'
    ws['E1'].value = 'Predominante (Gênero)'
    for c in 'ABCDE':
        ws[f'{c}1'].font = Font(bold=True)
    # UNIQUE list of categories
    ws['A2'].value = f"=UNIQUE({table_name}[{var_col}])"
    # Total por categoria
    ws['B2'].value = f"=IF(A2=\"\",\"\",ROWS(FILTER({table_name}[{gender_col}], {table_name}[{var_col}]=A2)))"
    # Simpson
    ws['C2'].value = (
        f"=IF(A2=\"\",\"\",LET(arr,FILTER({table_name}[{gender_col}],{table_name}[{var_col}]=A2),"
        f"u,UNIQUE(arr),p,COUNTIF(arr,u)/ROWS(arr),1-SUMPRODUCT(p^2)))"
    )
    # Shannon
    ws['D2'].value = (
        f"=IF(A2=\"\",\"\",LET(arr,FILTER({table_name}[{gender_col}],{table_name}[{var_col}]=A2),"
        f"u,UNIQUE(arr),p,COUNTIF(arr,u)/ROWS(arr),-SUMPRODUCT(p*LN(p))))"
    )
    # Predominante
    ws['E2'].value = (
        f"=IF(A2=\"\",\"\",LET(arr,FILTER({table_name}[{gender_col}],{table_name}[{var_col}]=A2),"
        f"u,UNIQUE(arr),c,COUNTIF(arr,u),INDEX(SORTBY(u,c,-1),1)))"
    )
    return ws


def build_race_sheet(wb, table_name, var_col, race_col):
    title = f"3C_RACA_{var_col}"
    ws = wb.create_sheet(title[:31])
    ws['A1'].value = 'Categoria'
    ws['B1'].value = 'Total'
    ws['C1'].value = 'Índice de Simpson (Raça)'
    ws['D1'].value = 'Índice de Shannon (Raça)'
    ws['E1'].value = 'Predominante (Raça)'
    for c in 'ABCDE':
        ws[f'{c}1'].font = Font(bold=True)
    ws['A2'].value = f"=UNIQUE({table_name}[{var_col}])"
    ws['B2'].value = f"=IF(A2=\"\",\"\",ROWS(FILTER({table_name}[{race_col}], {table_name}[{var_col}]=A2)))"
    ws['C2'].value = (
        f"=IF(A2=\"\",\"\",LET(arr,FILTER({table_name}[{race_col}],{table_name}[{var_col}]=A2),"
        f"u,UNIQUE(arr),p,COUNTIF(arr,u)/ROWS(arr),1-SUMPRODUCT(p^2)))"
    )
    ws['D2'].value = (
        f"=IF(A2=\"\",\"\",LET(arr,FILTER({table_name}[{race_col}],{table_name}[{var_col}]=A2),"
        f"u,UNIQUE(arr),p,COUNTIF(arr,u)/ROWS(arr),-SUMPRODUCT(p*LN(p))))"
    )
    ws['E2'].value = (
        f"=IF(A2=\"\",\"\",LET(arr,FILTER({table_name}[{race_col}],{table_name}[{var_col}]=A2),"
        f"u,UNIQUE(arr),c,COUNTIF(arr,u),INDEX(SORTBY(u,c,-1),1)))"
    )
    return ws


def apply_threshold_format(ws, idx_col_letter='C'):
    from openpyxl.formatting.rule import FormulaRule
    red = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    yellow = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    green = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    max_row = 2000
    last_col = get_column_letter(ws.max_column if ws.max_column > 5 else 5)
    ref_range = f"$A$2:${last_col}${max_row}"
    low, high = get_thresholds()
    ws.conditional_formatting.add(ref_range, FormulaRule(formula=[f"=${idx_col_letter}2<{low}"], fill=red))
    ws.conditional_formatting.add(ref_range, FormulaRule(formula=[f"=AND(${idx_col_letter}2>={low},${idx_col_letter}2<{high})"], fill=yellow))
    ws.conditional_formatting.add(ref_range, FormulaRule(formula=[f"=${idx_col_letter}2>={high}"], fill=green))


def sheet_name_for(prefix, var):
    name = f"{prefix}_{var}"
    return name[:31]


def build_gender_summary(wb, variables):
    ws = wb.create_sheet('3B_RESUMO_DIVERSIDADE_GENERO')
    headers = [
        'Variável', 'Categorias (Qtd)', 'Total (registros)',
        'Índice de Simpson (Gênero) — média ponderada',
        'Baixa (<0,6)', 'Moderada (0,6–<0,8)', 'Alta (>=0,8)', 'Sinal'
    ]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=1, column=j, value=h).font = Font(bold=True)
    for i, var in enumerate(variables, start=2):
        ws.cell(row=i, column=1, value=var)
        sname = sheet_name_for('3A_GENERO', var)
        # Define ranges
        cat = f"'{sname}'!A2:A1048576"
        tot = f"'{sname}'!B2:B1048576"
        sim = f"'{sname}'!C2:C1048576"
        # Categorias (Qtd)
        ws.cell(row=i, column=2, value=f"=IFERROR(COUNTA(FILTER({cat},{cat}<>\"\")),0)")
        # Total (registros)
        ws.cell(row=i, column=3, value=f"=IFERROR(SUM(FILTER({tot},{cat}<>\"\")),0)")
        # Simpson média ponderada
        ws.cell(row=i, column=4, value=(
            f"=IFERROR(SUMPRODUCT(FILTER({sim},{cat}<>\"\"),FILTER({tot},{cat}<>\"\"))/SUM(FILTER({tot},{cat}<>\"\")),\"\")"
        ))
        # Baixa/Moderada/Alta
        ws.cell(row=i, column=5, value=f"=LET(c,FILTER({sim},{cat}<>\"\"),IF(c=\"\",0,SUMPRODUCT(--(c<0.6))))")
        ws.cell(row=i, column=6, value=f"=LET(c,FILTER({sim},{cat}<>\"\"),IF(c=\"\",0,SUMPRODUCT(--(c>=0.6),--(c<0.8))))")
        ws.cell(row=i, column=7, value=f"=LET(c,FILTER({sim},{cat}<>\"\"),IF(c=\"\",0,SUMPRODUCT(--(c>=0.8))))")
        # Sinal por limiar, baseado na coluna D
        ws.cell(row=i, column=8, value=f"=IF(D{i}<0.6,\"🔴\",IF(D{i}<0.8,\"🟡\",\"🟢\"))")
    apply_threshold_format(ws, idx_col_letter='D')
    return ws


def build_race_summary(wb, variables):
    ws = wb.create_sheet('3D_RESUMO_DIVERSIDADE_RACA')
    headers = [
        'Variável', 'Categorias (Qtd)', 'Total (registros)',
        'Índice de Simpson (Raça) — média ponderada',
        'Baixa (<0,6)', 'Moderada (0,6–<0,8)', 'Alta (>=0,8)', 'Sinal'
    ]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=1, column=j, value=h).font = Font(bold=True)
    for i, var in enumerate(variables, start=2):
        ws.cell(row=i, column=1, value=var)
        sname = sheet_name_for('3C_RACA', var)
        cat = f"'{sname}'!A2:A1048576"
        tot = f"'{sname}'!B2:B1048576"
        sim = f"'{sname}'!C2:C1048576"
        ws.cell(row=i, column=2, value=f"=IFERROR(COUNTA(FILTER({cat},{cat}<>\"\")),0)")
        ws.cell(row=i, column=3, value=f"=IFERROR(SUM(FILTER({tot},{cat}<>\"\")),0)")
        ws.cell(row=i, column=4, value=(
            f"=IFERROR(SUMPRODUCT(FILTER({sim},{cat}<>\"\"),FILTER({tot},{cat}<>\"\"))/SUM(FILTER({tot},{cat}<>\"\")),\"\")"
        ))
        ws.cell(row=i, column=5, value=f"=LET(c,FILTER({sim},{cat}<>\"\"),IF(c=\"\",0,SUMPRODUCT(--(c<0.6))))")
        ws.cell(row=i, column=6, value=f"=LET(c,FILTER({sim},{cat}<>\"\"),IF(c=\"\",0,SUMPRODUCT(--(c>=0.6),--(c<0.8))))")
        ws.cell(row=i, column=7, value=f"=LET(c,FILTER({sim},{cat}<>\"\"),IF(c=\"\",0,SUMPRODUCT(--(c>=0.8))))")
        ws.cell(row=i, column=8, value=f"=IF(D{i}<0.6,\"🔴\",IF(D{i}<0.8,\"🟡\",\"🟢\"))")
    apply_threshold_format(ws, idx_col_letter='D')
    return ws


def main():
    if len(sys.argv) < 3:
        print("Usage: create_excel_only_report.py <input_csv> <output_xlsx>")
        sys.exit(1)
    csv_path = sys.argv[1]
    out_path = sys.argv[2]
    df = pd.read_csv(csv_path)

    wb = Workbook()
    ws_home = wb.active
    ws_home.title = '0_HOME'
    add_home(ws_home)

    # DADOS_BRUTOS
    ws_data = wb.create_sheet('DADOS_BRUTOS')
    # headers
    for j, col in enumerate(df.columns, start=1):
        ws_data.cell(row=1, column=j, value=col).font = Font(bold=True)
    # data
    for i, row in df.iterrows():
        for j, col in enumerate(df.columns, start=1):
            ws_data.cell(row=i+2, column=j, value=row[col])
    add_table(ws_data, max_row=len(df)+1, max_col=len(df.columns), name='TBL_DADOS')

    # identify gender & race columns
    gender_col = find_col(df, GENDER_NAMES)
    race_col = find_col(df, RACE_NAMES)
    exclude = set([c for c in [gender_col, race_col] if c])
    cat_cols = auto_categorical_cols(df, exclude)

    # Build per-variable sheets
    for var in cat_cols:
        if gender_col:
            ws = build_gender_sheet(wb, 'TBL_DADOS', var, gender_col)
            apply_threshold_format(ws, 'C')
        if race_col:
            ws = build_race_sheet(wb, 'TBL_DADOS', var, race_col)
            apply_threshold_format(ws, 'C')

    # Summaries
    if gender_col:
        build_gender_summary(wb, cat_cols)
    if race_col:
        build_race_summary(wb, cat_cols)

    # Visualizations
    build_visualizations(wb, cat_cols, bool(gender_col), bool(race_col))

    # Save workbook
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    wb.save(out_path)
    print(f"✓ Excel-only report created: {out_path}")


def build_visualizations(wb, variables, has_gender: bool, has_race: bool):
    ws = wb.create_sheet('5_VISUALIZACOES')
    ws['A1'].value = 'Visualizações (Índice de Simpson por categoria)'
    ws['A1'].font = Font(bold=True, size=14)
    row = 3
    for var in variables:
        if has_gender:
            sname = sheet_name_for('3A_GENERO', var)
            ws.cell(row=row, column=1, value=f"{var} — Gênero").font = Font(bold=True)
            chart = BarChart()
            chart.title = f"Índice de Simpson (Gênero) — {var}"
            chart.y_axis.title = 'Índice de Simpson'
            chart.x_axis.title = 'Categoria'
            # Categories: A2:A200; Data: C2:C200
            cats = Reference(wb[sname], min_col=1, min_row=2, max_row=200)
            data = Reference(wb[sname], min_col=3, min_row=1, max_row=200)  # include header for legend
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 10
            chart.width = 24
            ws.add_chart(chart, f"A{row+1}")
            row += 22
        if has_race:
            sname = sheet_name_for('3C_RACA', var)
            ws.cell(row=row, column=1, value=f"{var} — Raça").font = Font(bold=True)
            chart = BarChart()
            chart.title = f"Índice de Simpson (Raça) — {var}"
            chart.y_axis.title = 'Índice de Simpson'
            chart.x_axis.title = 'Categoria'
            cats = Reference(wb[sname], min_col=1, min_row=2, max_row=200)
            data = Reference(wb[sname], min_col=3, min_row=1, max_row=200)
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 10
            chart.width = 24
            ws.add_chart(chart, f"A{row+1}")
            row += 22

    # (saving handled in main)


if __name__ == '__main__':
    main()
