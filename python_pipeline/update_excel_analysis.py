#!/usr/bin/env python3
"""
Excel Analysis Updater
Updates the analysis sheets when new data is added to the DADOS_BRUTOS sheet
"""
 
import pandas as pd
import numpy as np
import seaborn as sns
from scipy import stats
from scipy.stats import chi2_contingency
import statsmodels.api as sm
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
import os
import tempfile
from datetime import datetime
import json
import unicodedata
from core.utils import load_config, get_thresholds, normalize, standardize_gender, standardize_race, find_gender_column, find_race_column
import warnings
warnings.filterwarnings('ignore')

class ExcelAnalysisUpdater:
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self.df = None
        self.analysis_results = {}
        self.original_columns = []
        
    def load_data_from_excel(self):
        """Load data from DADOS_BRUTOS sheet"""
        try:
            self.df = pd.read_excel(self.excel_path, sheet_name='DADOS_BRUTOS')
            self.original_columns = list(self.df.columns)
            print(f"✓ Dados carregados: {len(self.df)} registros, {len(self.df.columns)} colunas")
            return True
        except Exception as e:
            print(f"✗ Erro ao carregar dados: {e}")
            return False

    def _normalize(self, s: str) -> str:
        """Normalize a column name: lowercase, remove accents, non-alnum -> underscore"""
        try:
            s2 = unicodedata.normalize('NFKD', str(s)).encode('ASCII', 'ignore').decode('ASCII')
            return ''.join(ch if ch.isalnum() else '_' for ch in s2.lower()).strip('_')
        except Exception:
            return str(s).lower()

    def _ensure_age_bands(self):
        """Create a categorical 'faixa_etaria' from numeric age when applicable.
        - If a numeric age column exists (e.g., 'idade', 'age'), bin it.
        - If an existing faixa_etaria column is numeric-like, convert to bands.
        - If a categorical faixa_etaria exists, keep it and standardize to 'faixa_etaria'.
        """
        if self.df is None or self.df.empty:
            return
        norm_cols = {self._normalize(c): c for c in self.df.columns}

        # Candidate numeric age columns
        age_keys = [k for k in norm_cols if k == 'idade' or k.startswith('idade') or k == 'age']
        source_col = None
        age_series = None
        for k in age_keys:
            col = norm_cols[k]
            ser = pd.to_numeric(self.df[col], errors='coerce')
            if ser.notna().sum() > 0:
                source_col = col
                age_series = ser
                break

        # If not found, check an existing faixa_etaria-like column that might be numeric
        faixa_keys = [k for k in norm_cols if 'faixa' in k and 'etaria' in k]
        if source_col is None and faixa_keys:
            cand = norm_cols[faixa_keys[0]]
            ser = pd.to_numeric(self.df[cand], errors='coerce')
            if ser.notna().sum() > 0:
                source_col = cand
                age_series = ser

        if age_series is not None:
            bins = [-1, 17, 24, 34, 44, 54, 64, 200]
            labels = ['<18', '18-24', '25-34', '35-44', '45-54', '55-64', '65+']
            faixa = pd.cut(age_series, bins=bins, labels=labels)
            self.df['faixa_etaria'] = faixa.astype(object)
        else:
            # Standardize existing categorical faixa_etaria if present
            if faixa_keys and 'faixa_etaria' not in self.df.columns:
                self.df['faixa_etaria'] = self.df[norm_cols[faixa_keys[0]]]

    def _ensure_race_column(self):
        """Ensure a 'raca' column exists (creating an empty one or renaming a close match)."""
        if self.df is None:
            return
        norm_cols = {self._normalize(c): c for c in self.df.columns}
        race_keys = [k for k in norm_cols if k in ('raca', 'raca_cor', 'cor', 'race', 'etnia') or 'raca' in k or 'cor' in k or 'race' in k]
        if not race_keys:
            if 'raca' not in self.df.columns:
                self.df['raca'] = pd.Series([np.nan]*len(self.df))
        else:
            main = norm_cols[race_keys[0]]
            if 'raca' not in self.df.columns:
                try:
                    self.df.rename(columns={main: 'raca'}, inplace=True)
                except Exception:
                    pass
    
    def detect_data_types(self):
        """Automatically detect column types and characteristics"""
        data_info = {}
        
        for col in self.df.columns:
            col_data = self.df[col]
            unique_count = col_data.nunique()
            data_info[col] = {
                'type': 'categorical' if unique_count < 10 else 'numerical',
                'unique_values': unique_count,
                'null_count': col_data.isnull().sum(),
                'values': col_data.unique().tolist()[:10]
            }
        
        self.analysis_results['data_info'] = data_info
        return data_info

    def _find_gender_column(self):
        """Try to find the gender/sex column name in the dataset."""
        if self.df is None:
            return None
        candidates = []
        for col in self.df.columns:
            norm = self._normalize(col)
            if any(k in norm for k in ['genero', 'gnero', 'sexo', 'gender']):
                candidates.append(col)
        # Prefer exact matches first
        preferred = [c for c in candidates if self._normalize(c) in ['genero', 'gnero', 'sexo', 'gender']]
        return preferred[0] if preferred else (candidates[0] if candidates else None)

    def _standardize_gender(self, series: pd.Series) -> pd.Series:
        """Map various gender labels to standardized categories."""
        mapping = {
            'm': 'Masculino', 'masc': 'Masculino', 'masculino': 'Masculino', 'homem': 'Masculino', 'male': 'Masculino', 'man': 'Masculino',
            'f': 'Feminino', 'fem': 'Feminino', 'feminino': 'Feminino', 'mulher': 'Feminino', 'female': 'Feminino', 'woman': 'Feminino'
        }
        def norm_one(x):
            if pd.isna(x):
                return 'Outro/NS'
            s = self._normalize(x)
            return mapping.get(s, 'Masculino' if s in ['h'] else ('Feminino' if s in ['w'] else 'Outro/NS'))
        return series.apply(norm_one)

    def generate_gender_cross_diversity(self):
        """Compute diversity indices of gender distribution within each category of other variables."""
        results = []
        gender_col = self._find_gender_column()
        if gender_col is None:
            self.analysis_results['gender_cross_diversity'] = results
            return results
        gender_std = self._standardize_gender(self.df[gender_col])
        categorical_cols = [col for col in self.df.columns if col != gender_col and self.analysis_results['data_info'].get(col, {}).get('type') == 'categorical']
        for var in categorical_cols:
            for cat, sub in self.df.groupby(var):
                sub_gender = self._standardize_gender(sub[gender_col])
                counts = sub_gender.value_counts()
                total = counts.sum()
                if total == 0:
                    continue
                proportions = (counts / total).to_dict()
                # Compute indices
                simpson = 1 - sum((c/total)**2 for c in counts)
                shannon = -sum((c/total) * np.log(c/total) for c in counts if c > 0)
                # Predominant category
                pred_label = counts.idxmax()
                pred_pct = counts.max() / total
                results.append({
                    'Variavel': var,
                    'Categoria': cat,
                    'Total': int(total),
                    'Simpson': float(simpson),
                    'Shannon': float(shannon),
                    'Dominancia': float(pred_pct),
                    'Predominante': f"{pred_label} ({pred_pct*100:.1f}%)",
                    'Distribuicao': {k: f"{v/total*100:.1f}%" for k, v in counts.to_dict().items()}
                })
        self.analysis_results['gender_cross_diversity'] = results
        return results

    def _find_race_column(self):
        """Try to find the race/ethnicity column name in the dataset."""
        if self.df is None:
            return None
        candidates = []
        for col in self.df.columns:
            norm = self._normalize(col)
            if any(k in norm for k in ['raca', 'raça', 'cor', 'race', 'etnia', 'ethnic']):
                candidates.append(col)
        preferred = [c for c in candidates if self._normalize(c) in ['raca', 'cor', 'race', 'etnia']]
        return preferred[0] if preferred else (candidates[0] if candidates else None)

    def _standardize_race(self, series: pd.Series) -> pd.Series:
        """Map various race labels to standardized IBGE categories."""
        mapping = {
            'branca': 'Branca', 'branco': 'Branca',
            'preta': 'Preta', 'preto': 'Preta', 'negra': 'Preta', 'negro': 'Preta',
            'parda': 'Parda',
            'amarela': 'Amarela',
            'indigena': 'Indígena', 'indígena': 'Indígena',
            'nao_informado': 'Não informado', 'nao_declarado': 'Não informado', 'nd': 'Não informado', 'ns': 'Não informado', 'nr': 'Não informado'
        }
        def norm_one(x):
            if pd.isna(x):
                return 'Não informado'
            s = self._normalize(x)
            return mapping.get(s, 'Não informado')
        return series.apply(norm_one)

    def generate_race_cross_diversity(self):
        """Compute diversity indices of race distribution within each category of other variables."""
        results = []
        race_col = self._find_race_column()
        if race_col is None:
            self.analysis_results['race_cross_diversity'] = results
            return results
        categorical_cols = [col for col in self.df.columns if col != race_col and self.analysis_results['data_info'].get(col, {}).get('type') == 'categorical']
        for var in categorical_cols:
            for cat, sub in self.df.groupby(var):
                sub_race = self._standardize_race(sub[race_col])
                counts = sub_race.value_counts()
                total = counts.sum()
                if total == 0:
                    continue
                simpson = 1 - sum((c/total)**2 for c in counts)
                shannon = -sum((c/total) * np.log(c/total) for c in counts if c > 0)
                pred_label = counts.idxmax()
                pred_pct = counts.max() / total
                results.append({
                    'Variavel': var,
                    'Categoria': cat,
                    'Total': int(total),
                    'Simpson': float(simpson),
                    'Shannon': float(shannon),
                    'Dominancia': float(pred_pct),
                    'Predominante': f"{pred_label} ({pred_pct*100:.1f}%)",
                    'Distribuicao': {k: f"{v/total*100:.1f}%" for k, v in counts.to_dict().items()}
                })
        self.analysis_results['race_cross_diversity'] = results
        return results
    
    def generate_descriptive_stats(self):
        """Generate descriptive statistics for all columns"""
        stats_dict = {}
        
        for col in self.df.columns:
            col_data = self.df[col]
            if self.analysis_results['data_info'][col]['type'] == 'categorical':
                # Categorical statistics (robust to empty columns)
                value_counts = col_data.value_counts()
                percentages = col_data.value_counts(normalize=True) * 100
                if len(value_counts) > 0:
                    most_val = value_counts.index[0]
                    most_cnt = value_counts.iloc[0]
                    most_pct = percentages.iloc[0]
                else:
                    most_val = 'N/A'
                    most_cnt = 0
                    most_pct = 0.0
                stats_dict[col] = {
                    'type': 'categorical',
                    'value_counts': value_counts.to_dict(),
                    'percentages': percentages.to_dict(),
                    'most_common': most_val,
                    'most_common_count': most_cnt,
                    'most_common_pct': most_pct
                }
            else:
                # Numerical statistics
                stats_dict[col] = {
                    'type': 'numerical',
                    'count': len(col_data),
                    'mean': col_data.mean(),
                    'median': col_data.median(),
                    'std': col_data.std(),
                    'min': col_data.min(),
                    'max': col_data.max(),
                    'q25': col_data.quantile(0.25),
                    'q75': col_data.quantile(0.75)
                }
        
        self.analysis_results['descriptive_stats'] = stats_dict
        return stats_dict
    
    def perform_statistical_tests(self):
        """Perform statistical tests with user-friendly explanations"""
        tests_results = {}
        categorical_cols = [col for col in self.df.columns if self.analysis_results['data_info'][col]['type'] == 'categorical']
        
        # Chi-square tests for independence between categorical variables
        for i, col1 in enumerate(categorical_cols):
            for col2 in categorical_cols[i+1:]:
                contingency_table = pd.crosstab(self.df[col1], self.df[col2])
                
                if contingency_table.size > 0:
                    try:
                        chi2, p_value, dof, expected = chi2_contingency(contingency_table)
                        
                        # User-friendly interpretation
                        if p_value < 0.05:
                            interpretation = f"Há uma associação estatisticamente significativa entre {col1} e {col2} (p = {p_value:.4f}). Isso sugere que estas variáveis não são independentes."
                        else:
                            interpretation = f"Não há evidência suficiente para afirmar que existe uma associação entre {col1} e {col2} (p = {p_value:.4f}). As variáveis parecem ser independentes."
                        
                        tests_results[f"{col1}_vs_{col2}"] = {
                            'test_type': 'Qui-quadrado',
                            'chi2_statistic': chi2,
                            'p_value': p_value,
                            'degrees_of_freedom': dof,
                            'interpretation': interpretation,
                            'contingency_table': contingency_table.to_dict()
                        }
                    except Exception as e:
                        tests_results[f"{col1}_vs_{col2}"] = {
                            'test_type': 'Erro',
                            'error': str(e)
                        }
        
        self.analysis_results['statistical_tests'] = tests_results
        return tests_results
    
    def generate_diversity_indices(self):
        """Calculate diversity indices for categorical variables"""
        diversity_indices = {}
        
        for col in self.df.columns:
            if self.analysis_results['data_info'][col]['type'] == 'categorical':
                value_counts = self.df[col].value_counts()
                total = len(self.df)
                if len(value_counts) == 0 or total == 0:
                    continue
                
                # Simpson's Diversity Index
                simpson_index = 1 - sum((count/total)**2 for count in value_counts)
                
                # Shannon's Diversity Index
                proportions = value_counts / total
                shannon_index = -sum(p * np.log(p) for p in proportions if p > 0)
                
                # Dominance
                dominance = max(value_counts) / total
                
                diversity_indices[col] = {
                    'simpson_index': simpson_index,
                    'shannon_index': shannon_index,
                    'dominance': dominance,
                    'interpretation': self._interpret_diversity_index(simpson_index, col)
                }
        
        self.analysis_results['diversity_indices'] = diversity_indices
        return diversity_indices
    
    def _interpret_diversity_index(self, index, variable_name):
        """Provide user-friendly interpretation of diversity indices"""
        if index >= 0.8:
            return f"Alta diversidade em {variable_name} (índice = {index:.3f}). A distribuição é bem equilibrada entre diferentes categorias."
        elif index >= 0.6:
            return f"Diversidade moderada em {variable_name} (índice = {index:.3f}). Há uma boa distribuição, mas com algumas categorias predominantes."
        elif index >= 0.4:
            return f"Diversidade baixa em {variable_name} (índice = {index:.3f}). Algumas categorias são claramente predominantes."
        else:
            return f"Diversidade muito baixa em {variable_name} (índice = {index:.3f}). Há forte concentração em poucas categorias."
    
    def _get_second_most_common(self, value_counts):
        """Get the second most common value from value counts"""
        if len(value_counts) >= 2:
            # Sort by count and get second item
            sorted_items = sorted(value_counts.items(), key=lambda x: x[1], reverse=True)
            return f"{sorted_items[1][0]} ({sorted_items[1][1]})"
        return "N/A"
    
    def _get_distribution_summary(self, value_counts):
        """Get a summary of distribution pattern"""
        total = sum(value_counts.values())
        if len(value_counts) == 1:
            return "Concentrada em uma categoria"
        elif len(value_counts) == 2:
            return "Distribuição binária"
        elif len(value_counts) <= 4:
            return "Poucas categorias"
        else:
            return "Múltiplas categorias"
    
    def _get_mode(self, series):
        """Get mode of a series"""
        try:
            mode_result = series.mode()
            if len(mode_result) > 0:
                return mode_result.iloc[0]
            return "N/A"
        except:
            return "N/A"
    
    def create_excel_charts(self, workbook):
        """Create improved Excel charts with proper category display"""
        # Recreate visualization sheet to avoid duplicates
        if '5_VISUALIZACOES' in workbook.sheetnames:
            ws_old = workbook['5_VISUALIZACOES']
            workbook.remove(ws_old)
        chart_sheet = workbook.create_sheet('5_VISUALIZACOES')
        chart_sheet.cell(row=1, column=1, value="Gráficos de Análise").font = Font(bold=True, size=16)
        
        row_position = 3
        
        for col in self.df.columns:
            if self.analysis_results['data_info'][col]['type'] == 'categorical':
                # Create frequency table for the chart
                value_counts = self.df[col].value_counts()
                
                # Add data to chart sheet
                chart_sheet.cell(row=row_position, column=1, value=f"Gráfico de {col}:").font = Font(bold=True)
                
                # Create data table
                data_start_row = row_position + 2
                chart_sheet.cell(row=data_start_row, column=1, value="Categoria").font = Font(bold=True)
                chart_sheet.cell(row=data_start_row, column=2, value="Frequência").font = Font(bold=True)
                chart_sheet.cell(row=data_start_row, column=3, value="Percentagem").font = Font(bold=True)
                
                # Fill data
                total = len(self.df)
                for i, (category, count) in enumerate(value_counts.items()):
                    percentage = (count / total)
                    chart_sheet.cell(row=data_start_row + i + 1, column=1, value=category)
                    chart_sheet.cell(row=data_start_row + i + 1, column=2, value=int(count))
                    c = chart_sheet.cell(row=data_start_row + i + 1, column=3, value=float(percentage))
                    c.number_format = '0.0%'
                
                # Create Excel chart with improved formatting
                chart = BarChart()
                chart.type = "col"
                chart.style = 10
                
                # Set data range (include header so we can use titles_from_data=True)
                data = Reference(
                    worksheet=chart_sheet,
                    min_col=2,
                    min_row=data_start_row,
                    max_row=data_start_row + len(value_counts),
                    max_col=2,
                )
                
                # Set categories range (ONLY the category names, not including header)
                categories = Reference(worksheet=chart_sheet, 
                                     min_col=1, 
                                     min_row=data_start_row + 1, 
                                     max_row=data_start_row + len(value_counts))
                
                # Add data using header as series title
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(categories)
                
                # Configure chart titles and axes
                if col.lower() == 'gênero' or col.lower() == 'genero':
                    chart.title = 'Distribuição por Gênero'
                    chart.x_axis.title = 'Gênero'
                    chart.y_axis.title = 'Quantidade'
                elif col.lower() == 'cargo':
                    chart.title = 'Distribuição por Cargo'
                    chart.x_axis.title = 'Cargo'
                    chart.y_axis.title = 'Quantidade'
                elif col.lower() == 'faixa_etaria':
                    chart.title = 'Distribuição por Faixa Etária'
                    chart.x_axis.title = 'Faixa Etária'
                    chart.y_axis.title = 'Quantidade'
                elif col.lower() == 'departamento':
                    chart.title = 'Distribuição por Departamento'
                    chart.x_axis.title = 'Departamento'
                    chart.y_axis.title = 'Quantidade'
                elif col.lower() == 'instrução':
                    chart.title = 'Distribuição por Nível de Instrução'
                    chart.x_axis.title = 'Nível de Instrução'
                    chart.y_axis.title = 'Quantidade'
                elif col.lower() == 'tipo_contrato':
                    chart.title = 'Distribuição por Tipo de Contrato'
                    chart.x_axis.title = 'Tipo de Contrato'
                    chart.y_axis.title = 'Quantidade'
                else:
                    chart.title = f'Distribuição de {col}'
                    chart.x_axis.title = col
                    chart.y_axis.title = 'Quantidade'
                
                # Ensure category labels are visible
                chart.x_axis.tickLblPos = 'nextTo'
                try:
                    chart.x_axis.tickLblSkip = 1  # show every label
                except Exception:
                    pass
                
                # Add data labels
                chart.dataLabels = DataLabelList()
                chart.dataLabels.showVal = True
                chart.legend = None
                
                # Improve chart appearance
                chart.width = 12
                chart.height = 8
                
                # Position chart
                chart_position = f"E{data_start_row}"
                chart_sheet.add_chart(chart, chart_position)
                
                # Format data cells
                for i in range(len(value_counts)):
                    chart_sheet.cell(row=data_start_row + i + 1, column=2).alignment = Alignment(horizontal="center")
                    chart_sheet.cell(row=data_start_row + i + 1, column=3).alignment = Alignment(horizontal="center")
                
                # Update row position for next chart
                row_position = data_start_row + len(value_counts) + 10  # More space between charts
        
        # Format header row
        for cell in chart_sheet[1]:
            cell.font = Font(bold=True, size=16)
        
        return workbook

    def _auto_fit_columns(self, sheet, max_width=50):
        """Auto-ajusta larguras de colunas para melhor leitura"""
        try:
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if cell.value is not None:
                            max_length = max(max_length, len(str(cell.value)))
                    except Exception:
                        pass
                sheet.column_dimensions[column_letter].width = min(max_length + 2, max_width)
        except Exception:
            pass

    def _ensure_home_sheet(self, workbook):
        """Cria/atualiza a aba 0_HOME com instruções e navegação"""
        name = '0_HOME'
        if name in workbook.sheetnames:
            ws = workbook[name]
            # Limpa conteúdo
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.value = None
        else:
            ws = workbook.create_sheet(name, 0)

        ws.cell(row=1, column=1, value='Análise de Diversidade - Hub').font = Font(bold=True, size=18)
        ws.cell(row=2, column=1, value=f"Última atualização: {datetime.now().strftime('%d/%m/%Y %H:%M')}")

        ws.cell(row=4, column=1, value='Como usar').font = Font(bold=True, size=14)
        ws.cell(row=5, column=1, value='1. Atualize os dados na aba DADOS_BRUTOS (não apague cabeçalhos).')
        ws.cell(row=6, column=1, value='2. Execute o atualizador para regerar as análises.')
        ws.cell(row=7, column=1, value='3. Use os links abaixo para navegar entre os resultados.')

        ws.cell(row=9, column=1, value='Navegação rápida').font = Font(bold=True, size=14)
        targets = [
            'DADOS_BRUTOS', '1_VISAO_GERAL',
            '2_TESTES_ESTATISTICOS', '3_INDICES_DIVERSIDADE',
            '3A_DIVERSIDADE_GENERO', '3B_RESUMO_DIVERSIDADE_GENERO',
            '3C_DIVERSIDADE_RACA', '3D_RESUMO_DIVERSIDADE_RACA',
            '5_VISUALIZACOES'
        ]
        r = 10
        for t in targets:
            if t in workbook.sheetnames:
                c = ws.cell(row=r, column=1, value=f'Ir para {t}')
                c.hyperlink = f"#'{t}'!A1"
                c.style = 'Hyperlink'
                r += 1

        ws.cell(row=10, column=3, value='Dicas').font = Font(bold=True, size=12)
        ws.cell(row=11, column=3, value='- Use os filtros nas tabelas para explorar os dados.')
        ws.cell(row=12, column=3, value='- Valores destacados indicam possíveis inconsistências (validação de dados).')
        ws.cell(row=13, column=3, value='- Consulte AJUDA_GLOSSARIO para entender os indicadores.')

        # Legend for diversity colors
        base_row = 15
        ws.cell(row=base_row, column=1, value='Legenda de cores (Índice de Simpson)').font = Font(bold=True, size=12)
        # Red: baixa
        ws.cell(row=base_row+1, column=1, value='< 0,6 = Baixa diversidade').fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        # Yellow: moderada
        ws.cell(row=base_row+2, column=1, value='0,6 – <0,8 = Diversidade moderada').fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        # Green: alta
        ws.cell(row=base_row+3, column=1, value='≥ 0,8 = Alta diversidade').fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        # Short interpretation text
        ws.cell(row=base_row+5, column=1, value='Interpretação (resumo):').font = Font(bold=True)
        ws.cell(row=base_row+6, column=1, value='• Índice de Simpson mede equilíbrio entre grupos (0 a 1).')
        ws.cell(row=base_row+7, column=1, value='• Quanto mais próximo de 1, mais distribuída é a composição por grupo.')
        ws.cell(row=base_row+8, column=1, value='• Sinalização aplica-se a gênero e raça nas abas 3A/3B/3C/3D.')

        # Executive KPIs
        def _safe_pct(a, b):
            try:
                return (a / b) if b else 0.0
            except Exception:
                return 0.0

        # Compute KPIs using standardized gender and race
        # 1) % Mulheres total
        pct_mulheres = None
        genero_col = None
        try:
            genero_col = self._find_gender_column()
            if genero_col is not None:
                gser = self._standardize_gender(self.df[genero_col])
                den = gser.notna().sum()
                num = (gser == 'Feminino').sum()
                pct_mulheres = _safe_pct(num, den)
            else:
                pct_mulheres = 0.0
        except Exception:
            pct_mulheres = 0.0

        # 2) % Mulheres em liderança (heurística)
        # Leadership detection using configurable keywords/columns
        def _leadership_mask(df):
            return self._leadership_mask(df)

        pct_mulheres_lider = 0.0
        try:
            if genero_col is not None:
                gser = self._standardize_gender(self.df[genero_col])
                mlead = _leadership_mask(self.df)
                den = (gser.notna() & mlead).sum()
                num = ((gser == 'Feminino') & mlead).sum()
                pct_mulheres_lider = _safe_pct(num, den)
        except Exception:
            pct_mulheres_lider = 0.0

        # 3) % PPI (pretos/pardos) total
        pct_ppi = 0.0
        try:
            race_col = self._find_race_column()
            if race_col is not None:
                rser = self._standardize_race(self.df[race_col])
                den = rser.notna().sum()
                num = ((rser == 'Preta') | (rser == 'Parda')).sum()
                pct_ppi = _safe_pct(num, den)
        except Exception:
            pct_ppi = 0.0

        # 4) Índice Simpson geral (gênero) e (raça) + semáforo
        def _simpson_from_counts(counts):
            try:
                total = sum(counts.values())
                if not total:
                    return 0.0
                return 1 - sum((c/total)**2 for c in counts.values())
            except Exception:
                return 0.0

        idx_gen = 0.0
        idx_rac = 0.0
        try:
            if genero_col is not None:
                gser = self._standardize_gender(self.df[genero_col])
                idx_gen = _simpson_from_counts(gser.value_counts().to_dict())
        except Exception:
            idx_gen = 0.0
        try:
            race_col = self._find_race_column()
            if race_col is not None:
                rser = self._standardize_race(self.df[race_col])
                idx_rac = _simpson_from_counts(rser.value_counts().to_dict())
        except Exception:
            idx_rac = 0.0

        def _semaforo(v):
            return '🔴' if v < 0.6 else ('🟡' if v < 0.8 else '🟢')

        kpi_row = base_row + 11
        ws.cell(row=kpi_row, column=1, value='Painel Executivo').font = Font(bold=True, size=14)
        # Labels
        ws.cell(row=kpi_row+1, column=1, value='% Mulheres (total)')
        ws.cell(row=kpi_row+2, column=1, value='% Mulheres em liderança')
        ws.cell(row=kpi_row+3, column=1, value='% PPI (pretos/pardos)')
        ws.cell(row=kpi_row+4, column=1, value='Índice Simpson (Gênero)')
        ws.cell(row=kpi_row+4, column=2, value=_semaforo(idx_gen))
        ws.cell(row=kpi_row+5, column=1, value='Índice Simpson (Raça)')
        ws.cell(row=kpi_row+5, column=2, value=_semaforo(idx_rac))
        # Values
        v1 = ws.cell(row=kpi_row+1, column=2, value=float(pct_mulheres)); v1.number_format = '0.0%'
        v2 = ws.cell(row=kpi_row+2, column=2, value=float(pct_mulheres_lider)); v2.number_format = '0.0%'
        v3 = ws.cell(row=kpi_row+3, column=2, value=float(pct_ppi)); v3.number_format = '0.0%'
        v4 = ws.cell(row=kpi_row+4, column=3, value=float(idx_gen)); v4.number_format = '0.000'
        v5 = ws.cell(row=kpi_row+5, column=3, value=float(idx_rac)); v5.number_format = '0.000'

        self._auto_fit_columns(ws)
        return workbook

    def _load_leadership_config(self):
        """Load leadership detection config if available."""
        candidates = [
            os.environ.get('DIVERSITY_CONFIG'),
            os.path.join(os.getcwd(), 'config_diversidade.json'),
            os.path.join(os.getcwd(), 'leadership_config.json'),
        ]
        for path in candidates:
            if path and os.path.exists(path):
                try:
                    with open(path, 'r', encoding='utf-8') as f:
                        return json.load(f)
                except Exception:
                    pass
        return {}

    def _leadership_mask(self, df: pd.DataFrame) -> pd.Series:
        """Return boolean Series indicating leadership rows based on config and heuristics."""
        try:
            cfg = self._load_leadership_config()
            col_hints = cfg.get('leadership_column_hints', [
                'cargo', 'funcao', 'função', 'job', 'title', 'posição', 'posicao', 'nivel', 'nível', 'senior', 'senioridade', 'lead', 'gestao', 'gestão', 'role', 'position', 'level', 'seniority'
            ])
            explicit_cols = cfg.get('leadership_columns', [])
            keywords = cfg.get('leadership_keywords', [
                'gerent', 'diretor', 'coordenador', 'supervisor', 'lider', 'líder', 'head', 'chief', 'c-level', 'vp', 'presidente', 'gestor', 'chef',
                'manager', 'director', 'lead', 'supervisor', 'coordinator', 'head', 'chief', 'vp', 'president'
            ])
            cols = []
            # Match by hints
            for c in df.columns:
                n = self._normalize(c)
                if any(h in n for h in [self._normalize(hh) for hh in col_hints]):
                    cols.append(c)
            # Add explicit columns if exist
            for c in explicit_cols:
                if c in df.columns and c not in cols:
                    cols.append(c)
            if not cols:
                return pd.Series([False]*len(df))
            keys_norm = [self._normalize(k) for k in keywords]
            mask = pd.Series([False]*len(df))
            for c in cols:
                vals = df[c].astype(str).str.lower()
                m = vals.apply(lambda x: any(kn in x for kn in keys_norm))
                mask = mask | m
            return mask
        except Exception:
            return pd.Series([False]*len(df))

    def _apply_low_diversity_formatting(self, workbook):
        """Apply conditional formatting to highlight low diversity rows in 3A/3B/3C/3D."""
        targets = [
            ('3A_DIVERSIDADE_GENERO', 'Índice de Simpson (Gênero)'),
            ('3B_RESUMO_DIVERSIDADE_GENERO', 'Índice de Simpson (Gênero) — média ponderada'),
            ('3C_DIVERSIDADE_RACA', 'Índice de Simpson (Raça)'),
            ('3D_RESUMO_DIVERSIDADE_RACA', 'Índice de Simpson (Raça) — média ponderada'),
        ]
        red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        yellow_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
        green_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        for sheet_name, header_text in targets:
            if sheet_name in workbook.sheetnames:
                sh = workbook[sheet_name]
                if sh.max_row < 2:
                    continue
                # Find target column by header
                target_col_idx = None
                for j, cell in enumerate(sh[1], start=1):
                    if str(cell.value).strip() == header_text:
                        target_col_idx = j
                        break
                if not target_col_idx:
                    continue
                target_col_letter = get_column_letter(target_col_idx)
                last_col_letter = get_column_letter(sh.max_column)
                # Number format for target column
                for r in range(2, sh.max_row + 1):
                    sh[f"{target_col_letter}{r}"].number_format = '0.000'
                ref_range = f"$A$2:${last_col_letter}${sh.max_row}"
                low, high = get_thresholds()
                rule_red = FormulaRule(formula=[f"=${target_col_letter}2<{low}"], fill=red_fill)
                sh.conditional_formatting.add(ref_range, rule_red)
                rule_yellow = FormulaRule(formula=[f"=AND(${target_col_letter}2>={low},${target_col_letter}2<{high})"], fill=yellow_fill)
                sh.conditional_formatting.add(ref_range, rule_yellow)
                rule_green = FormulaRule(formula=[f"=${target_col_letter}2>={high}"], fill=green_fill)
                sh.conditional_formatting.add(ref_range, rule_green)
        return workbook

    def _ensure_glossary_sheet(self, workbook):
        """Cria/atualiza a aba AJUDA_GLOSSARIO com explicações"""
        name = 'AJUDA_GLOSSARIO'
        if name in workbook.sheetnames:
            ws = workbook[name]
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.value = None
        else:
            ws = workbook.create_sheet(name)

        ws.cell(row=1, column=1, value='Glossário e Metodologia').font = Font(bold=True, size=16)
        ws.cell(row=3, column=1, value='Índice de Simpson: 1 - Σ p_i^2 (quanto mais perto de 1, mais diversidade).')
        ws.cell(row=4, column=1, value='Índice de Shannon: -Σ p_i ln(p_i) (maior valor indica maior diversidade).')
        ws.cell(row=5, column=1, value='Qui-quadrado: testa associação entre variáveis categóricas (p<0,05 indica associação).')
        ws.cell(row=7, column=1, value='Boas práticas: mantenha categorias consistentes (use listas suspensas); evite células vazias.')
        ws.cell(row=8, column=1, value='Limitações: resultados dependem da qualidade dos dados e tamanho de amostra.')
        self._auto_fit_columns(ws)
        return workbook

    def _ensure_lists_sheet_and_validations(self, workbook):
        """Cria LISTAS e aplica validação de dados na aba DADOS_BRUTOS"""
        name = 'LISTAS'
        if name in workbook.sheetnames:
            ws_list = workbook[name]
            for row in ws_list.iter_rows(min_row=1, max_row=ws_list.max_row, min_col=1, max_col=ws_list.max_column):
                for cell in row:
                    cell.value = None
        else:
            ws_list = workbook.create_sheet(name)

        # Colunas categóricas
        cat_cols = [c for c in self.df.columns if self.analysis_results['data_info'][c]['type'] == 'categorical']
        list_ranges = {}
        col_idx = 1
        for col in cat_cols:
            values = (
                self.df[col]
                .dropna()
                .astype(str)
                .drop_duplicates()
                .sort_values()
                .tolist()
            )
            # Default lists for key fields when empty
            col_norm = self._normalize(col)
            if len(values) == 0 and (col_norm in ['raca', 'cor', 'raca_cor', 'race', 'etnia'] or 'raca' in col_norm or 'cor' in col_norm):
                values = ['Branca', 'Preta', 'Parda', 'Amarela', 'Indígena', 'Não informado']
            if len(values) == 0 and (col_norm in ['genero', 'sexo', 'gender'] or 'genero' in col_norm or 'sexo' in col_norm):
                values = ['Feminino', 'Masculino', 'Outro/NS']
            col_letter = get_column_letter(col_idx)
            ws_list.cell(row=1, column=col_idx, value=col).font = Font(bold=True)
            for i, v in enumerate(values, start=2):
                ws_list.cell(row=i, column=col_idx, value=v)
            end_row = 1 + max(1, len(values))
            list_ranges[col] = (col_letter, 2, end_row)
            col_idx += 1

        self._auto_fit_columns(ws_list)

        # Aplicar validação na aba DADOS_BRUTOS
        if 'DADOS_BRUTOS' not in workbook.sheetnames:
            return workbook

        ws_raw = workbook['DADOS_BRUTOS']
        headers = [cell.value for cell in ws_raw[1]]
        max_row = ws_raw.max_row
        for j, header in enumerate(headers, start=1):
            if header in list_ranges:
                col_letter = get_column_letter(j)
                list_col_letter, start_row, end_row = list_ranges[header]
                dv = DataValidation(
                    type='list',
                    formula1=f"={name}!${list_col_letter}${start_row}:${list_col_letter}${end_row}",
                    allow_blank=True,
                    showErrorMessage=True,
                    errorTitle='Valor inválido',
                    error='Selecione um valor da lista.'
                )
                ws_raw.add_data_validation(dv)
                dv.ranges.add(f"${col_letter}$2:${col_letter}${max_row}")

                # Destaque valores fora da lista
                formula = f"=COUNTIF({name}!${list_col_letter}:${list_col_letter},${col_letter}2)=0"
                red_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                rule = FormulaRule(formula=[formula], fill=red_fill)
                ws_raw.conditional_formatting.add(f"${col_letter}$2:${col_letter}${max_row}", rule)

        return workbook

    def _ensure_raw_table_and_freeze(self, workbook):
        """Converte DADOS_BRUTOS em Tabela e congela cabeçalhos; congela nas abas de análise."""
        if 'DADOS_BRUTOS' in workbook.sheetnames:
            ws = workbook['DADOS_BRUTOS']
            ws.freeze_panes = 'A2'
            try:
                if not ws.tables:
                    end_cell = f"{get_column_letter(ws.max_column)}{ws.max_row}"
                    tbl = Table(displayName='TBL_DADOS', ref=f"A1:{end_cell}")
                    style = TableStyleInfo(name='TableStyleMedium2', showFirstColumn=False,
                                           showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                    tbl.tableStyleInfo = style
                    ws.add_table(tbl)
            except Exception:
                pass

        # Congelar cabeçalhos nas demais abas
        for sname in workbook.sheetnames:
            if sname not in ('0_HOME', 'AJUDA_GLOSSARIO'):
                try:
                    ws = workbook[sname]
                    if ws.max_row >= 2:
                        ws.freeze_panes = 'A2'
                except Exception:
                    pass
        return workbook

    def _add_header_comments(self, workbook):
        """Adiciona comentários de ajuda em cabeçalhos selecionados."""
        # Estatísticas descritivas
        s1 = '2_ESTATISTICAS_DESCRITIVAS'
        if s1 in workbook.sheetnames:
            ws = workbook[s1]
            headers = [c.value for c in ws[1]]
            help_text = {
                'Média': 'Média aritmética dos valores numéricos.',
                'Mediana': 'Valor central que divide a distribuição em duas metades.',
                'Desvio Padrão': 'Mede a dispersão em relação à média.',
                'Variância': 'Quadrado do desvio padrão.',
                'Q1 (25%)': 'Primeiro quartil (25% dos dados abaixo).',
                'Q3 (75%)': 'Terceiro quartil (75% dos dados abaixo).',
                'IQR': 'Intervalo interquartil (Q3 - Q1).',
                'Coeficiente de Variação': 'Desvio padrão / média (em %).',
                'Percentagem': 'Frequência relativa do valor mais comum.'
            }
            for j, h in enumerate(headers, start=1):
                if h in help_text and ws.cell(row=1, column=j).comment is None:
                    ws.cell(row=1, column=j).comment = Comment(help_text[h], 'Sistema')

        # Índices de diversidade
        s2 = '3_INDICES_DIVERSIDADE'
        if s2 in workbook.sheetnames:
            ws = workbook[s2]
            headers = [c.value for c in ws[1]]
            help_text = {
                'Índice de Simpson': '1 - Σ p_i^2. Quanto mais próximo de 1, maior a diversidade (geral).',
                'Índice de Shannon': '-Σ p_i ln(p_i). Valores maiores indicam mais diversidade.',
                'Dominância': 'Proporção da categoria mais frequente.'
            }
            for j, h in enumerate(headers, start=1):
                if h in help_text and ws.cell(row=1, column=j).comment is None:
                    ws.cell(row=1, column=j).comment = Comment(help_text[h], 'Sistema')
        return workbook

    def _interpret_index_label(self, index: float, escopo: str) -> str:
        try:
            if index >= 0.8:
                return f'Alta diversidade de {escopo} (índice = {index:.3f}).'
            elif index >= 0.6:
                return f'Diversidade moderada de {escopo} (índice = {index:.3f}).'
            else:
                return f'Baixa diversidade de {escopo} (índice = {index:.3f}).'
        except Exception:
            return 'N/A'
    
    def update_analysis_sheets(self):
        """Update all analysis sheets while preserving the DADOS_BRUTOS sheet and VBA (if .xlsm)"""
        # Load existing workbook; preserve VBA for .xlsm
        keep_vba = self.excel_path.lower().endswith('.xlsm')
        book = load_workbook(self.excel_path, keep_vba=keep_vba)

        # Remove existing analysis sheets (keep only DADOS_BRUTOS)
        for sheet_name in list(book.sheetnames):
            if sheet_name != 'DADOS_BRUTOS':
                ws = book[sheet_name]
                book.remove(ws)

        # Persist cleanup to file
        book.save(self.excel_path)

        # Create new analysis sheets using a writer that preserves VBA if needed
        with pd.ExcelWriter(
            self.excel_path,
            engine='openpyxl',
            mode='a',
            if_sheet_exists='replace',
            engine_kwargs={'keep_vba': keep_vba}
        ) as writer:
            # Replace DADOS_BRUTOS to include newly created columns like 'raca'
            df_out = self.df.copy()
            # Drop derived 'faixa_etaria' if it wasn't originally present
            if 'faixa_etaria' in df_out.columns and 'faixa_etaria' not in self.original_columns:
                df_out = df_out.drop(columns=['faixa_etaria'])
            # Ensure 'raca' exists
            if 'raca' not in df_out.columns:
                df_out['raca'] = np.nan
            df_out.to_excel(writer, sheet_name='DADOS_BRUTOS', index=False)
            # 1. Data Overview
            overview_data = []
            for col, info in self.analysis_results['data_info'].items():
                overview_data.append({
                    'Variável': col,
                    'Tipo': info['type'],
                    'Valores Únicos': info['unique_values'],
                    'Valores Nulos': info['null_count'],
                    'Exemplo de Valores': ', '.join(map(str, info['values'][:5]))
                })
            
            overview_df = pd.DataFrame(overview_data)
            overview_df.to_excel(writer, sheet_name='1_VISAO_GERAL', index=False)
            
            # (Sheet 2 removed by request)
            
            # 2. Statistical Tests
            tests_data = []
            for test_name, result in self.analysis_results['statistical_tests'].items():
                if result['test_type'] == 'Qui-quadrado':
                    tests_data.append({
                        'Teste': test_name,
                        'Tipo': 'Qui-quadrado',
                        'Estatística': f"{result['chi2_statistic']:.3f}",
                        'Valor-p': f"{result['p_value']:.4f}",
                        'Interpretação': result['interpretation']
                    })
            
            if tests_data:
                tests_df = pd.DataFrame(tests_data)
                tests_df.to_excel(writer, sheet_name='2_TESTES_ESTATISTICOS', index=False)
            
            # 3. Diversity Indices (overall per variable)
            diversity_data = []
            # Build weighted Simpson by variable for Gender and Race from cross results
            gender_cross = self.analysis_results.get('gender_cross_diversity', [])
            race_cross = self.analysis_results.get('race_cross_diversity', [])
            gender_w = {}
            race_w = {}
            if gender_cross:
                tmp = {}
                for rcd in gender_cross:
                    tmp.setdefault(rcd['Variavel'], []).append((rcd['Simpson'], rcd['Total']))
                for var, vals in tmp.items():
                    tw = sum(w for _, w in vals) or 1
                    gender_w[var] = sum(s*w for s, w in vals) / tw
            if race_cross:
                tmp = {}
                for rcd in race_cross:
                    tmp.setdefault(rcd['Variavel'], []).append((rcd['Simpson'], rcd['Total']))
                for var, vals in tmp.items():
                    tw = sum(w for _, w in vals) or 1
                    race_w[var] = sum(s*w for s, w in vals) / tw
            for col, indices in self.analysis_results['diversity_indices'].items():
                diversity_data.append({
                    'Variável': col,
                    'Índice de Simpson': f"{indices['simpson_index']:.3f}",
                    'Índice de Shannon': f"{indices['shannon_index']:.3f}",
                    'Dominância': f"{indices['dominance']:.3f}",
                    'Interpretação (Geral)': indices['interpretation'],
                    'Índice de Simpson (Gênero) — média ponderada': gender_w.get(col, float('nan')) if gender_w else float('nan'),
                    'Interpretação (Gênero)': self._interpret_index_label(gender_w.get(col, float('nan')), 'gênero') if gender_w else 'N/A',
                    'Índice de Simpson (Raça) — média ponderada': race_w.get(col, float('nan')) if race_w else float('nan'),
                    'Interpretação (Raça)': self._interpret_index_label(race_w.get(col, float('nan')), 'raça') if race_w else 'N/A',
                })
            
            diversity_df = pd.DataFrame(diversity_data)
            diversity_df.to_excel(writer, sheet_name='3_INDICES_DIVERSIDADE', index=False)

            # 3A. Gender Cross-Diversity (gender within each category of other variables)
            cross = self.analysis_results.get('gender_cross_diversity', [])
            if cross:
                cross_rows = []
                # Determine known gender labels present overall to create stable columns
                labels_set = set()
                for r in cross:
                    labels_set.update([k for k in r['Distribuicao'].keys()])
                label_cols = []
                for lbl in ['Feminino', 'Masculino', 'Outro/NS']:
                    if lbl in labels_set:
                        label_cols.append(lbl)
                # Add any other labels found
                for lbl in sorted(labels_set):
                    if lbl not in label_cols:
                        label_cols.append(lbl)
                for r in cross:
                    row = {
                        'Variável': r['Variavel'],
                        'Categoria da Variável': r['Categoria'],
                        'Total': r['Total'],
                        'Índice de Simpson (Gênero)': r['Simpson'],
                        'Índice de Shannon (Gênero)': r['Shannon'],
                        'Dominância (Gênero)': r['Dominancia'],
                        'Predominante (Gênero)': r['Predominante']
                    }
                    for lbl in label_cols:
                        row[f"{lbl} (%)"] = r['Distribuicao'].get(lbl, '0.0%')
                    cross_rows.append(row)
                pd.DataFrame(cross_rows).to_excel(writer, sheet_name='3A_DIVERSIDADE_GENERO', index=False)

                # 4B. Summary per variable (weighted averages and counts by diversity bands)
                from collections import defaultdict
                by_var = defaultdict(list)
                for r in cross:
                    by_var[r['Variavel']].append(r)
                summary_rows = []
                for var, rows in by_var.items():
                    total_weight = sum(x['Total'] for x in rows) or 1
                    w_simpson = sum(x['Simpson'] * x['Total'] for x in rows) / total_weight
                    w_dom = sum(x['Dominancia'] * x['Total'] for x in rows) / total_weight
                    low = sum(1 for x in rows if x['Simpson'] < 0.6)
                    mid = sum(1 for x in rows if 0.6 <= x['Simpson'] < 0.8)
                    high = sum(1 for x in rows if x['Simpson'] >= 0.8)
                    # critical category (lowest Simpson)
                    worst = min(rows, key=lambda x: x['Simpson'])
                    summary_rows.append({
                        'Variável': var,
                        'Categorias (Qtd)': len(rows),
                        'Total (registros)': total_weight,
                        'Índice de Simpson (Gênero) — média ponderada': w_simpson,
                        'Dominância (Gênero) — média ponderada': w_dom,
                        'Baixa (<0,6)': low,
                        'Moderada (0,6–<0,8)': mid,
                        'Alta (>=0,8)': high,
                        'Sinal': '🔴' if w_simpson < 0.6 else ('🟡' if w_simpson < 0.8 else '🟢'),
                        'Categoria mais crítica': worst['Categoria'],
                        'Predominante na crítica (Gênero)': worst['Predominante'],
                        'Simpson crítico': worst['Simpson']
                    })
                pd.DataFrame(summary_rows).to_excel(writer, sheet_name='3B_RESUMO_DIVERSIDADE_GENERO', index=False)

            # 3C. Race Cross-Diversity
            cross_r = self.analysis_results.get('race_cross_diversity', [])
            if cross_r:
                cross_rows_r = []
                labels_set_r = set()
                for r in cross_r:
                    labels_set_r.update([k for k in r['Distribuicao'].keys()])
                label_cols_r = []
                for lbl in ['Branca', 'Preta', 'Parda', 'Amarela', 'Indígena', 'Não informado']:
                    if lbl in labels_set_r:
                        label_cols_r.append(lbl)
                for lbl in sorted(labels_set_r):
                    if lbl not in label_cols_r:
                        label_cols_r.append(lbl)
                for r in cross_r:
                    row = {
                        'Variável': r['Variavel'],
                        'Categoria da Variável': r['Categoria'],
                        'Total': r['Total'],
                        'Índice de Simpson (Raça)': r['Simpson'],
                        'Índice de Shannon (Raça)': r['Shannon'],
                        'Dominância (Raça)': r['Dominancia'],
                        'Predominante (Raça)': r['Predominante']
                    }
                    for lbl in label_cols_r:
                        row[f"{lbl} (%)"] = r['Distribuicao'].get(lbl, '0.0%')
                    cross_rows_r.append(row)
                pd.DataFrame(cross_rows_r).to_excel(writer, sheet_name='3C_DIVERSIDADE_RACA', index=False)

                # 3D. Summary per variable (Race)
                from collections import defaultdict
                by_var_r = defaultdict(list)
                for r in cross_r:
                    by_var_r[r['Variavel']].append(r)
                summary_rows_r = []
                for var, rows in by_var_r.items():
                    total_weight = sum(x['Total'] for x in rows) or 1
                    w_simpson = sum(x['Simpson'] * x['Total'] for x in rows) / total_weight
                    w_dom = sum(x['Dominancia'] * x['Total'] for x in rows) / total_weight
                    low = sum(1 for x in rows if x['Simpson'] < 0.6)
                    mid = sum(1 for x in rows if 0.6 <= x['Simpson'] < 0.8)
                    high = sum(1 for x in rows if x['Simpson'] >= 0.8)
                    worst = min(rows, key=lambda x: x['Simpson'])
                    summary_rows_r.append({
                        'Variável': var,
                        'Categorias (Qtd)': len(rows),
                        'Total (registros)': total_weight,
                        'Índice de Simpson (Raça) — média ponderada': w_simpson,
                        'Dominância (Raça) — média ponderada': w_dom,
                        'Baixa (<0,6)': low,
                        'Moderada (0,6–<0,8)': mid,
                        'Alta (>=0,8)': high,
                        'Sinal': '🔴' if w_simpson < 0.6 else ('🟡' if w_simpson < 0.8 else '🟢'),
                        'Categoria mais crítica': worst['Categoria'],
                        'Predominante na crítica (Raça)': worst['Predominante'],
                        'Simpson crítico': worst['Simpson']
                    })
                pd.DataFrame(summary_rows_r).to_excel(writer, sheet_name='3D_RESUMO_DIVERSIDADE_RACA', index=False)
            
            # 4. Detailed Tables (skip legacy/aux columns)
            skip_cols_norm = { 'faixa_etaria_texto', 'idade_min', 'idade_max', 'idade' }
            for col in self.df.columns:
                if self.analysis_results['data_info'][col]['type'] == 'categorical':
                    if self._normalize(col) in skip_cols_norm:
                        continue
                    value_counts = self.df[col].value_counts()
                    percentages = (self.df[col].value_counts(normalize=True) * 100).round(1)
                    
                    freq_table = pd.DataFrame({
                        'Valor': value_counts.index,
                        'Frequência': value_counts.values,
                        'Percentagem': percentages.values
                    })
                    
                    # Add percentage symbol
                    freq_table['Percentagem'] = freq_table['Percentagem'].apply(lambda x: f'{x:.1f}%')
                    freq_table.to_excel(writer, sheet_name=f'4_TABELA_{col.upper()}', index=False)
            
            # Format all sheets
            for sheet_name in writer.book.sheetnames:
                if sheet_name != 'DADOS_BRUTOS':  # Don't format the raw data sheet
                    sheet = writer.book[sheet_name]
                    
                    # Format header row
                    for cell in sheet[1]:
                        cell.font = Font(bold=True)
                        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        cell.font = Font(color="FFFFFF", bold=True)
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Auto-adjust column widths
                    for column in sheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        sheet.column_dimensions[column_letter].width = adjusted_width

            # Apply conditional formatting to 3B and 3D
            self._apply_low_diversity_formatting(writer.book)
        
        # Now add charts to the final workbook
        print("📊 Criando gráficos no Excel...")
        final_book = load_workbook(self.excel_path, keep_vba=keep_vba)
        final_book = self.create_excel_charts(final_book)
        # HOME, GLOSSÁRIO, LISTAS, validação, tabela, congelamento e comentários
        final_book = self._ensure_home_sheet(final_book)
        final_book = self._ensure_glossary_sheet(final_book)
        final_book = self._ensure_lists_sheet_and_validations(final_book)
        final_book = self._ensure_raw_table_and_freeze(final_book)
        final_book = self._add_header_comments(final_book)
        final_book.save(self.excel_path)
        print(f"✓ Gráficos, navegação, validação e ajuda adicionados ao Excel")
        
        print(f"✓ Análises atualizadas: {self.excel_path}")
        return self.excel_path
    
    def run_update(self):
        """Run the complete update process"""
        if not self.load_data_from_excel():
            return False
        
        print("📊 Analisando estrutura dos dados...")
        self._ensure_age_bands()
        self._ensure_race_column()
        self.detect_data_types()
        
        print("📈 Gerando estatísticas descritivas...")
        self.generate_descriptive_stats()
        
        print("🔬 Realizando testes estatísticos...")
        self.perform_statistical_tests()
        
        print("🎯 Calculando índices de diversidade...")
        self.generate_diversity_indices()
        # Also compute gender cross-diversity
        print("👥 Calculando diversidade de gênero por variável...")
        self.generate_gender_cross_diversity()
        # Also compute race cross-diversity
        print("🌈 Calculando diversidade de raça por variável...")
        self.generate_race_cross_diversity()
        
        print("📋 Atualizando abas de análise...")
        self.update_analysis_sheets()
        
        print("✅ Atualização concluída!")
        return True

def main():
    """Main function to run the update"""
    import sys
    
    if len(sys.argv) < 2:
        print("Uso: python update_excel_analysis.py <caminho_excel>")
        sys.exit(1)
    
    excel_path = sys.argv[1]
    
    updater = ExcelAnalysisUpdater(excel_path)
    updater.run_update()

if __name__ == "__main__":
    main()
